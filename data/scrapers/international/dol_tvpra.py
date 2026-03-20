"""DOL TVPRA List of Goods scraper (Excel download).

Downloads the Trafficking Victims Protection Reauthorization Act (TVPRA)
List of Goods Produced by Child Labor or Forced Labor. Maintained by
DOL ILAB, this list identifies 15 goods from Pakistan produced using
child or forced labor (e.g., bricks, carpets, coal, cotton, glass bangles,
leather, surgical instruments, textiles, wheat).

Source: https://www.dol.gov/agencies/ilab/reports/child-labor/list-of-goods
Schedule: Annually (0 5 1 3 *)
Priority: P1 — Official US government list of tainted goods
"""

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import logging
import re

from data.scrapers.base_api_scraper import BaseAPIScraper

logger = logging.getLogger(__name__)

TVPRA_PAGE_URL: str = (
    "https://www.dol.gov/agencies/ilab/reports/child-labor/list-of-goods"
)
PAKISTAN_COUNTRY: str = "Pakistan"


class DOLTVPRAScraper(BaseAPIScraper):
    """Scraper for DOL TVPRA List of Goods (Pakistan rows).

    Fetches the TVPRA page, locates the Excel download link,
    downloads the file, parses with openpyxl, and filters for
    Pakistan entries.
    """

    name: str = "dol_tvpra"
    source_url: str = TVPRA_PAGE_URL
    schedule: str = "0 5 1 3 *"
    priority: str = "P1"
    rate_limit_delay: float = 2.0
    request_timeout: float = 60.0

    async def scrape(self) -> list[dict[str, Any]]:
        """Download TVPRA Excel file, parse, and filter for Pakistan."""
        logger.info("[%s] Fetching TVPRA page to find Excel URL", self.name)

        excel_url = await self._find_excel_url()
        if not excel_url:
            logger.warning("[%s] Could not locate Excel download URL", self.name)
            return []

        excel_path = await self._download_excel(excel_url)
        if excel_path is None:
            return []

        rows = self._parse_excel(excel_path)
        logger.info(
            "[%s] Parsed %d total rows from TVPRA Excel", self.name, len(rows)
        )

        pak_rows = [
            row for row in rows
            if PAKISTAN_COUNTRY.lower() in str(row.get("country", "")).lower()
        ]
        logger.info(
            "[%s] Filtered to %d Pakistan rows", self.name, len(pak_rows)
        )

        return [self._to_record(row) for row in pak_rows]

    async def _find_excel_url(self) -> str | None:
        """Scrape the TVPRA page to find the Excel file download link."""
        from bs4 import BeautifulSoup

        response = await self.fetch(TVPRA_PAGE_URL)
        soup = BeautifulSoup(response.text, "html.parser")

        for link in soup.find_all("a", href=True):
            href = link["href"]
            text = link.get_text(strip=True).lower()
            if any(ext in href.lower() for ext in [".xlsx", ".xls"]):
                return (
                    href if href.startswith("http")
                    else f"https://www.dol.gov{href}"
                )
            if "excel" in text or "download" in text and "list" in text:
                candidate = href
                if any(ext in candidate.lower() for ext in [".xlsx", ".xls"]):
                    return (
                        candidate if candidate.startswith("http")
                        else f"https://www.dol.gov{candidate}"
                    )

        # Fallback: known URL pattern
        fallback_url = (
            "https://www.dol.gov/sites/dolgov/files/ILAB/"
            "ListofGoods.xlsx"
        )
        logger.warning(
            "[%s] Using fallback Excel URL: %s", self.name, fallback_url
        )
        return fallback_url

    async def _download_excel(self, url: str) -> Path | None:
        """Download the Excel file to local storage."""
        raw_dir = self.get_raw_dir()
        filename = url.split("/")[-1].split("?")[0]
        if not filename.endswith((".xlsx", ".xls")):
            filename = f"tvpra_list_{self.run_id}.xlsx"
        file_path = raw_dir / filename

        try:
            content = await self.fetch_bytes(url)
            file_path.write_bytes(content)
            logger.info(
                "[%s] Downloaded Excel: %s (%d bytes)",
                self.name, file_path, len(content),
            )
            return file_path
        except Exception as exc:
            logger.error("[%s] Excel download failed: %s", self.name, exc)
            return None

    @staticmethod
    def _parse_excel(path: Path) -> list[dict[str, Any]]:
        """Parse the TVPRA Excel file using openpyxl.

        Returns a list of dicts, one per row, with normalized header keys.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl is required; install with: pip install openpyxl")
            return []

        rows: list[dict[str, Any]] = []
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                logger.error("No active worksheet in %s", path)
                return []

            # Read headers from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [
                str(cell).strip().lower().replace(" ", "_") if cell else f"col_{i}"
                for i, cell in enumerate(header_row)
            ]

            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                row_dict: dict[str, Any] = {}
                for i, cell in enumerate(row_cells):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    row_dict[key] = cell
                if any(v is not None for v in row_dict.values()):
                    rows.append(row_dict)

            wb.close()
        except Exception as exc:
            logger.error("Failed to parse Excel %s: %s", path, exc)

        return rows

    def _to_record(self, row: dict[str, Any]) -> dict[str, Any]:
        """Convert a TVPRA Excel row to a statistical_reports record."""
        good = str(row.get("good", row.get("product", ""))).strip()
        labor_type = str(
            row.get("child_labor", row.get("cl", ""))
        ).strip()
        forced_labor = str(
            row.get("forced_labor", row.get("fl", ""))
        ).strip()

        exploitation_types: list[str] = []
        if labor_type and labor_type.lower() not in ("", "none", "nan"):
            exploitation_types.append("child_labor")
        if forced_labor and forced_labor.lower() not in ("", "none", "nan"):
            exploitation_types.append("forced_labor")

        # Try to extract year from any date-like field
        year = self._extract_year_from_row(row)

        return {
            "source_name": self.name,
            "report_year": year,
            "report_title": "TVPRA List of Goods Produced by Child/Forced Labor",
            "indicator": f"tvpra_listed_good:{good}",
            "value": 1,
            "unit": "listed_good",
            "geographic_scope": "Pakistan",
            "pdf_url": None,
            "extraction_method": "excel_openpyxl",
            "extraction_confidence": 0.95,
            "victim_gender": None,
            "victim_age_bracket": None,
            "good_name": good,
            "exploitation_types": exploitation_types,
            "scraped_at": datetime.now(timezone.utc).isoformat(),
        }

    @staticmethod
    def _extract_year_from_row(row: dict[str, Any]) -> str:
        """Try to find a year value in any cell of the row."""
        for value in row.values():
            val_str = str(value) if value is not None else ""
            match = re.search(r"20[0-2]\d", val_str)
            if match:
                return match.group()
        return str(datetime.now(timezone.utc).year)

    def validate(self, record: dict[str, Any]) -> bool:
        """Validate a TVPRA record.

        Requires source_name, indicator, and that the good_name is non-empty.
        """
        if not record.get("source_name"):
            return False
        if not record.get("indicator"):
            return False
        good = record.get("good_name", "")
        if not good or str(good).lower() in ("none", "nan", ""):
            return False
        return True
